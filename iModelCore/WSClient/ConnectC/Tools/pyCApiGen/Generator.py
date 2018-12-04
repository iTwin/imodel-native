#!Python
#--------------------------------------------------------------------------------------
#
#     $Source: ConnectC/Tools/pyCApiGen/Generator.py $
#
#  $Copyright: (c) 2017 Bentley Systems, Incorporated. All rights reserved. $
#
#--------------------------------------------------------------------------------------
import sys
from xml.dom import minidom
import os
import optparse
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module='openpyxl')
from openpyxl import load_workbook


from SchemaWriter.Helpers.Api import Api
from SchemaWriter.Helpers.ECSchema import ECSchema
from SchemaWriter.HeaderWriters.SchemaApiPublicHeaderWriter import CallStatus
from SchemaWriter.HeaderWriters.SchemaApiPublicHeaderWriter import SchemaApiPublicHeaderWriter
from SchemaWriter.SourceWriters.SchemaApiSourceWriter import SchemaApiSourceWriter
from SchemaWriter.HeaderWriters.SchemaBufferHeaderWriter import SchemaBufferHeaderWriter
from SchemaWriter.HeaderWriters.SchemaBufferPublicHeaderWriter import SchemaBufferPublicHeaderWriter
from SchemaWriter.SourceWriters.SchemaBufferSourceWriter import SchemaBufferSourceWriter
from BufferHeaderWriter import BufferHeaderWriter
from BufferSourceWriter import BufferSourceWriter

#############################################
# ------------- Status Codes -------------- #
#############################################
status_codes = {
    'ERROR400': CallStatus(400, 'The response from the server contained a 400, Bad Request, http error'),
    'ERROR401': CallStatus(401, 'The response from the server contained a 401, Unauthorized, http error'),
    'ERROR403': CallStatus(403, 'The response from the server contained a 403, Forbidden, http error'),
    'ERROR404': CallStatus(404, 'The response from the server contained a 404, Not Found, http error'),
    'ERROR409': CallStatus(409, 'The response from the server contained a 409, Conflict, http error'),
    'ERROR500': CallStatus(500, 'The response from the server contained a 500, Internal Server, http error'),
    'SUCCESS': CallStatus(0, 'Successful operation'),
    'INVALID_PARAMETER': CallStatus(-100, 'Invalid parameter passed to function'),
    'PROPERTY_HAS_NOT_BEEN_SET': CallStatus(-101, 'The buffer property passed to function has not been set in the buffer'),
    'INTERNAL_MEMORY_ERROR': CallStatus(-102, 'Memory failed to initialize interally.'),
    'LOGIN_FAILED': CallStatus(-200, 'The login attempt was not completed successfully.'),
    'SSL_REQUIRED': CallStatus(-201, 'SSL is required for communication with the server.'),
    'NOT_ENOUGH_RIGHTS': CallStatus(-202, 'The user does not have the proper rights.'),
    'REPOSITORY_NOT_FOUND': CallStatus(-204, 'The repository you attempted to query was not found.'),
    'SCHEMA_NOT_FOUND': CallStatus(-205, 'The schema you attempted to query was not found.'),
    'CLASS_NOT_FOUND': CallStatus(-206, 'The class you attempted to query was not found.'),
    'PROPERTY_NOT_FOUND': CallStatus(-207, 'The property you attempted to query was not found.'),
    'INSTANCE_NOT_FOUND': CallStatus(-210, 'The instance you attempted to query was not found.'),
    'FILE_NOT_FOUND': CallStatus(-211, 'The file you attempted to query was not found.'),
    'NOT_SUPPORTED': CallStatus(-212, 'The action you attempted is not supported.'),
    'NO_SERVER_LICENSE': CallStatus(-213, 'A server license was not found.'),
    'NO_CLIENT_LICENSE': CallStatus(-214, 'A client license was not found.'),
    'TO_MANY_BAD_LOGIN_ATTEMPTS': CallStatus(-215, 'To many unsuccessful login attempts have happened.'),
}

#-------------------------------------------------------------------------------------------
# Parse the schema file
# @bsimethod                                                        
#-------------------------------------------------------------------------------------------
def ParseSchemaFile(api, excelFile):
    workbook = load_workbook(filename=excelFile, read_only=True)
    worksheet = workbook.get_sheet_by_name('autoGenClasses')
    unique_ecschemas = {}
    for row in worksheet.iter_rows(range_string=worksheet.calculate_dimension(), row_offset=2):
        if row[0].value is None or row[1].value is None or row[2].value is None:
            break
        if row[0].value + row[1].value + row[2].value + row[3].value not in unique_ecschemas:
            unique_ecschemas[row[0].value + row[1].value + row[2].value + row[3].value] = \
                ECSchema(row[0].value, row[1].value, row[2].value, row[3].value, api, status_codes)
        unique_ecschemas[row[0].value + row[1].value + row[2].value + row[3].value].add_ecclass(row)

    #Read Relationships sheet
    #TODO: Move this code to a separate method. Maybe rework how how info loaded from sheet.
    rel_worksheet = workbook.get_sheet_by_name('Relationships')
    unique__rel_defs = {}
    for row in rel_worksheet.iter_rows(range_string=rel_worksheet.calculate_dimension(), row_offset=2):
        if row[0].value is None or row[1].value is None or row[2].value is None:
            break
        if row[0].value + row[1].value + row[2].value + row[3].value not in unique_ecschemas:
            unique_ecschemas[row[0].value + row[1].value + row[2].value + row[3].value] = \
                ECSchema(row[0].value, row[1].value, row[2].value, row[3].value, api, status_codes)
        unique_ecschemas[row[0].value + row[1].value + row[2].value + row[3].value].add_ecrelationship_specification(
                                            row[4].value, row[5].value, row[6].value, row[7].value, row[8].value, row[9].value)

    for ecschema in unique_ecschemas.values():
        filename = ecschema.get_filename()
        if not os.path.isabs(filename):
            filename = os.path.join(os.path.dirname(os.path.realpath(__file__)), filename)
        #load the schema xml
        xmldoc = minidom.parse(filename)
        ecschemasXml = xmldoc.getElementsByTagName('ECSchema')
        for ecschemaXml in ecschemasXml:
            if ecschemaXml.attributes['schemaName'].value == ecschema.get_name():
                if ecschemaXml.attributes['xmlns'].value != 'http://www.bentley.com/schemas/Bentley.ECXML.2.0':
                    raise IOError("{0} file's xmlns is not support. This tool only supports the "
                                  "'http://www.bentley.com/schemas/Bentley.ECXML.2.0' format"
                                  .format(ecschema.get_name()))
                ecschema.init_xml(ecschemaXml)
                break
    return unique_ecschemas

#-------------------------------------------------------------------------------------------
# Schema Auto Generation
# @bsimethod                                                        
#-------------------------------------------------------------------------------------------
def GenerateSchemaCode(srcDir, publicApiDir, api, schemas):
    for ecschema in schemas.values():        
        # SCHEMA-LEVEL-API
        apiPublicHeaderFile = publicApiDir + '{0}/{1}GenPublic.h' .format(ecschema.get_url_descriptor(), ecschema.get_name())
        schemaApiPublicHeaderWriter = SchemaApiPublicHeaderWriter(ecschema, apiPublicHeaderFile, api, status_codes)
        schemaApiPublicHeaderWriter.write_header()
        schemaApiSourceWriter = SchemaApiSourceWriter(ecschema, srcDir + '{0}/{1}Gen.cpp' .format(ecschema.get_url_descriptor(), ecschema.get_name()), api, status_codes)
        schemaApiSourceWriter.write_source()

        # SCHEMA-LEVEL - BUFFER
        schemaBufferHeaderFile = srcDir + '{0}/{1}BufferGen.h' .format(ecschema.get_url_descriptor(), ecschema.get_name())
        schemaBufferHeaderWriter = SchemaBufferHeaderWriter(ecschema, schemaBufferHeaderFile, api, status_codes)
        schemaBufferHeaderWriter.write_header()

        schemaBufferPublicHeaderFile = publicApiDir + '{0}/{1}BufferGenPublic.h' .format(ecschema.get_url_descriptor(), ecschema.get_name())
        schemaBufferPublicHeaderWriter = SchemaBufferPublicHeaderWriter(ecschema, schemaBufferPublicHeaderFile, api, status_codes)
        schemaBufferPublicHeaderWriter.write_header()

        schemaBufferSourceFile = srcDir + '{0}/{1}BufferGen.cpp' .format(ecschema.get_url_descriptor(), ecschema.get_name())            
        schemaBufferSourceWriter = SchemaBufferSourceWriter(ecschema, schemaBufferSourceFile, api, status_codes)
        schemaBufferSourceWriter.write_source()

    # API-LEVEL    
    apiBufferHeaderFile =  srcDir + '{0}BufferGen.h'.format(api.get_upper_api_acronym())
    bufferHeaderWriter = BufferHeaderWriter(schemas.values(), apiBufferHeaderFile, api, status_codes)
    bufferHeaderWriter.write_header()
    
    apiBufferSourceFile = srcDir + '{0}BufferGen.cpp'.format(api.get_upper_api_acronym())
    bufferSourceWriter = BufferSourceWriter(schemas.values(), apiBufferSourceFile, api, status_codes)
    bufferSourceWriter.write_source()

#-------------------------------------------------------------------------------------------
# @bsimethod                                                        Robert.Priest 2/2017
#-------------------------------------------------------------------------------------------
def main():
    #defaults
    defaultTargetSourceDir = os.path.join(os.path.dirname(os.path.realpath(__file__)), '../../')
    defaultTargetPublicApiDir = os.path.join(os.path.dirname(os.path.realpath(__file__)), '../../../PublicApi/WebServices/ConnectC/')
    defaultExcelFile = os.path.join(os.path.dirname(os.path.realpath(__file__)), 'autoGenClasses.xlsx')

    #arguments
    parser = optparse.OptionParser()
    parser.add_option ("-a", "--apiAcronym", dest="acronym", action="store", help="set api acronym", default="CWSCC")
    parser.add_option ("-n", "--apiName", dest="apiname", action="store", help="set api name", default="ConnectWebServicesClientC")
    parser.add_option ("-f", "--excelGenFile", dest="excelgenfile", action="store", help="excel auto generation description file", default=defaultExcelFile)
    parser.add_option ("-s", "--targetsourcedir", dest="targetsourcedir", action="store", help="target directory for source files", default=defaultTargetSourceDir)    
    parser.add_option ("-p", "--targetpublicapidir", dest="targetpublicapidir", action="store", help="target directory for public api files", default=defaultTargetPublicApiDir)    
    (options, args) = parser.parse_args()
    
    #parse the schema file and the excel sheet
    schemas = None
    try:
        print 'Parsing schema file "{0}"...'.format(options.excelgenfile)
        api = Api(options.acronym, options.apiname)
        schemas = ParseSchemaFile(api, options.excelgenfile)
    except Exception, ex:
        print("{0}".format (ex))
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)
        return
    
    #generate the schema code.
    try:
        print 'Generating schema code...'
        GenerateSchemaCode(options.targetsourcedir, options.targetpublicapidir,api, schemas)
    except Exception, ex:
        print("{0}".format (ex))
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)
        return

    print "Done"

#-------------------------------------------------------------------------------------------
# @bsimethod                                                        Robert.Priest 2/2017
#-------------------------------------------------------------------------------------------
if __name__ == '__main__':
    main()
    